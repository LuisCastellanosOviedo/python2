import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]


products_per_supplier = {}
total_value_per_supplier = {}

print(type(product_list))
print(product_list.max_row)

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value

    # Calculation number of products per supplier
    if supplier_name in products_per_supplier:
        products_per_supplier[supplier_name] = products_per_supplier.get(supplier_name) + 1
    else:
        print("Adding a new supplier !!!!!")
        products_per_supplier[supplier_name] = 1


    # Calculation total value of inventory per supplier
    if supplier_name in 
    total_value_per_supplier.[supplier_name] = inventory * price

print(f" final dictionary ======>>>> {products_per_supplier}")






